import os

import nsepy
import openpyxl
import pytest

import datetime
from datetime import timedelta

f_list = []

os.chdir('/storage/emulated/0/Download')


def read_data():
    wb = openpyxl.load_workbook('monthly_recom.xlsx')
    # wb = openpyxl.load_workbook('quote'+str(datetime.date.today())+'.xlsx')
    ws = wb['Sheet2']
    row = ws.max_row
    l = []
    for i in range(1, row + 1):
        l1 = []
        stock_name = ws.cell(i, 1)
        stock_type = ws.cell(i, 2)
        # stock_price=ws.cell(i,3)
        l1.insert(0, stock_name.value)
        l1.insert(1, stock_type.value)
        # l1.insert(2,stock_price.value)
        l.insert(i - 1, l1)

    print(l)
    return l


@pytest.mark.parametrize('data', read_data())
def test_create_chart(data):
    buy = []
    date_y = datetime.date.today() - timedelta(days=4)
    df = nsepy.get_history(symbol=data[0], start=datetime.date(2020, 5, 1), end=datetime.date.today())
    if float(df[['High']].max()) > float(data[1]):
        buy.append(data[0])
        print('Buy Price acgieved' + data[0])
        buy.append(data[1])
        buy.append(float(df[['High']].max()))
        profit = round(float(df[['High']].max()) - float(data[1]), 2)
        print('Profit is:', profit)
        buy.append(profit)

        pp = (round(float(df[['High']].max()) - float(data[1]), 2) / float(data[1])) * 100
        buy.append(round(pp, 2))
        print('Profit percentage is:', round(pp))
        f_list.append(buy)


# print(f_list)

def test_toExcel():
    book = openpyxl.load_workbook('monthly_recom.xlsx')

    sheet = book.create_sheet(str(datetime.date.today().strftime('%b')) + '_profits')
    for row in f_list:
        sheet.append(row)

    book.save('monthly_recom.xlsx')
