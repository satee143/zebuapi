import os

import nsepy
import openpyxl
import pytest

import datetime
from datetime import timedelta

global df, buy_list
buy_list = []
os.chdir('/storage/emulated/0/Download')


def read_data():
    wb = openpyxl.load_workbook('quote_list.xlsx')
    ws = wb['quote1']
    row = ws.max_row
    l = []

    for i in range(1, row + 1):
        l1 = []
        stock_name = ws.cell(i, 1)

        l.insert(0, stock_name.value)

    print(l)
    return l


@pytest.mark.parametrize('data', read_data())
def test_create_chart(data):
    date_y = datetime.date.today() - timedelta(days=4)
    df = nsepy.get_history(symbol=data, start=date_y, end=datetime.date.today())
    buy = []
    if (float(df.iloc[[-1]]['Open']) >= float(df.iloc[[-1]]['High'])):
        buy.append(data)
        buy.append('Sell')
        print('Sell Recommande Price for ' + data + ':', float(df.iloc[[-1]]['Low']))
        buy.append(float(df.iloc[[-1]]['Low']))
        buy_list.append(buy)

    elif (float(df.iloc[[-1]]['Open']) <= float(df.iloc[[-1]]['Low'])):
        buy.append(data)
        buy.append('Buy')
        print('Buy Recommanded Price for ' + data + ':', float(df.iloc[[-1]]['Open']))
        buy.append(float(df.iloc[[-1]]['High']))
        buy_list.append(buy)


def test_add():
    book = openpyxl.load_workbook('quote.xlsx')
    ws = book.create_sheet(str(datetime.date.today()))

    for row in buy_list:
        ws.append(row)

    book.save('quote.xlsx', )
    read_data()
