import os

import nsepy
import openpyxl
import plotly.graph_objects as go

import datetime
from datetime import timedelta

global df, buy_list
global updates
buy_list = []
updates = []
import pytest

data = 'INFY'
buy = []
os.chdir('/storage/emulated/0/Download')


def read_data():
    wb = openpyxl.load_workbook('nse_quotes.xlsx')
    ws = wb['data2']
    row = ws.max_row
    l = []

    for i in range(1, row + 1):
        l1 = []
        stock_name = ws.cell(i, 1)

        l.insert(0, stock_name.value)

    print(l)
    return l


@pytest.mark.parametrize('data', read_data())
def test_create_chart(data):
    buy = []
    date_y = datetime.date.today() - timedelta(days=30)
    df = nsepy.get_history(symbol='TATASTEEL', start=date_y, end=datetime.date.today())
    df.reset_index('Date', inplace=True)
    df['Date'] = df['Date'].astype(str)
    df['Date'] = df['Date'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d'))
    df.set_index('Date', inplace=True)
    ohlc_dict = {'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last'}

    df = df.resample('W', how=ohlc_dict).dropna(how='any')
    cols = ['Open', 'High', 'Low', 'Close']
    df = df[cols]

    # df = df['Close'].resample('1M').ohlc()
    df.reset_index('Date', inplace=True)

    df['HA_Close'] = ((df['Open'] + df['High'] + df['Low'] + df['Close']) / 4)

    df['HA_Open'] = (df['Open'].shift(1) + df['Open'].shift(1)) / 2
    df.iloc[0, df.columns.get_loc("HA_Open")] = (df.iloc[0]['Open'] + df.iloc[0]['Close']) / 2
    df['HA_High'] = df[['High', 'Low', 'HA_Open', 'HA_Close']].max(axis=1)
    df['HA_Low'] = ((df[['High', 'Low', 'HA_Open', 'HA_Close']].min(axis=1)))

    # print(df[['HA_Open','HA_High','HA_Low','HA_Close']])
    # df.reset_index('Date', inplace=True)
    # fig = go.Figure(data=[go.Ohlc(x=df['Date'], Open=df['Open'],High=df['High'], Low=df['Low'], Close=df['Close'])])

    # fig = go.Figure(data=[go.Candlestick(x=df['Date'], open=df['HA_Open'],high=df['HA_High'], low=df['HA_Low'], close=df['HA_Close'])])

    if (float(df.iloc[[-2]]['HA_Close']) <= float(df.iloc[[-2]]['HA_Open'])):
        if (float(df.iloc[[-1]]['HA_Close']) >= float(df.iloc[[-1]]['HA_Open'])):
            if (float(df.iloc[[-1]]['HA_High']) > float(df.iloc[[-2]]['HA_High']) and float(
                    df.iloc[[-1]]['HA_Low']) < float(df.iloc[[-2]]['HA_Low'])):
                print('Buy Recommanded Price for ' + data + ' :', round(float(df.iloc[[-1]]['HA_High']), 2))
                os.chdir('/storage/emulated/0/recom')
                df[['Date', 'HA_Open', 'HA_High', 'HA_Low', 'HA_Close']].to_excel(data + '.xlsx')
                fig = go.Figure(data=[
                    go.Candlestick(x=df['Date'], open=df['HA_Open'], high=df['HA_High'], low=df['HA_Low'],
                                   close=df['HA_Close'])])
                fig.write_html(data + '.html')
                buy.append(data)
                buy.append('buy')
                buy.append(round(float(df.iloc[[-1]]['HA_High']), 2))
                buy_list.append(buy)
    elif (float(df.iloc[[-2]]['HA_Close']) >= float(df.iloc[[-2]]['HA_Open'])):
        if ((float(df.iloc[[-1]]['HA_Close']) <= float(df.iloc[[-1]]['HA_Open']))):
            print('Sell Recommanded Price for ' + data + ':', round(float(df.iloc[[-1]]['HA_Low']), 2))


def test_add():
    os.chdir('/storage/emulated/0/Download')

    book = openpyxl.load_workbook('monthly_recom.xlsx')
    ws = book.create_sheet(str(datetime.date.today().strftime('%b')) + 'Week_Recom')
    for row in buy_list:
        ws.append(row)

    book.save('monthly_recom.xlsx')
