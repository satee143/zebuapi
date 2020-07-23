import os

os.chdir('/storage/emulated/0/ds')
print(os.getcwd())
'''for lis in os.listdir():
	print(lis)
from openpyxl import Workbook

workbook = Workbook()
spreadsheet = workbook.active'''

from openpyxl import load_workbook

workbook = load_workbook(filename="abcd.xlsx")
# print(workbook.sheetnames)
sheet = workbook['Sheet1']
'''
#spreadsheet = workbook.active
#spreadsheet
#Worksheet "Sheet 1">
l=[]
for column in sheet.iter_cols(min_row=2,max_row=sheet.max_row, min_col=1,  max_col=sheet.max_column,values_only=True):
	print(column)
	l.append([colun.value for colun in column ])
for x in l:
	for y in x:
		print(y)
	print()'''
#	for (colum) in column:
# print(colum)
# print(column)
'''for x in sheet[1:2]:
	#for y in x:
	print(x.value)
print(sheet['A2'].value)
print(sheet.cell(1,1).value)

#spreadsheet["A1"] = "Hello"
#spreadsheet["B1"] = "World!"
'''
from openpyxl.styles import Font, Alignment, Side, colors

Bold_Font = Font(size=20)
Big_Red_Text = Font(color=colors.RED, size=20)
Center_Aligned_Text = Alignment(horizontal="center")
Double_Border_Side = Side(border_style="double")
'''Square_Border = Border(top=double_border_side,right=double_border_side, 
bottom=double_border_side, 
left=double_border_side) '''

sheet["A2"].font = Bold_Font
sheet["A3"].font = Big_Red_Text
sheet["A4"].alignment = Center_Aligned_Text
# sheet["A5"].border = Square_Border
workbook.save(filename="abcd.xlsx")
# workbook.save(filename="HelloWorld.xlsx")'''
