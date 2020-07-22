import os

import nsepy
import openpyxl
import pytest

import datetime
from datetime import timedelta

global df, buy_list
buy_list = []
os.chdir('/storage/emulated/0/Download')


def read_data():
    wb = openpyxl.load_workbook('ab.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    l = []

    for i in range(1, row + 1):
        l1 = []
        stock_name = ws.cell(i, 1)
        stock_type = ws.cell(i, 2)
        stock_price = ws.cell(i, 3)

        l1.insert(0, stock_name.value)
        l1.insert(1, stock_type.value)
        l1.insert(2, stock_price.value)

        l.insert(i - 1, l1)

    return l


@pytest.mark.parametrize('data', read_data())
def test_create_chart(data):
    date_y = datetime.date.today() - timedelta(days=2)
    df = nsepy.get_history(symbol=data[1], start=date_y, end=datetime.date.today())
    print(df[['Open', 'High', 'Low']])
    if data[1] == 'Buy':
        if df.iloc[[-1]]['High'] >= data[2]:
            print(data[2])
            print(data[0], 'Price achieved')
            print('Day High', df.iloc[[-1]]['High'])


    elif data[1] == 'Sell':
        if df.iloc[[-1]]['Low'] <= data[2]:
            print(data[2])
            print(data[0], 'Sell Price achieved')
            print('Day Low', df.iloc[[-1]]['Low'])


'''		
	buy=[]
	if (float(df.iloc[[-1]]['Open']) >= float(df.iloc[[-1]]['High'])):
		buy.append(data)
		buy.append('Sell')
		print('Sell Recommande Price for '+ data + ':',float(df.iloc[[-1]]['Low']))
		buy.append(float(df.iloc[[-1]]['Low']))
		buy_list.append(buy)
	
	elif (float(df.iloc[[-1]]['Open']) <= float(df.iloc[[-1]]['Low'])):
		buy.append(data)
		buy.append('Buy')
		print('Buy Recommanded Price for '+ data + ':',float(df.iloc[[-1]]['Open']))
		buy.append(float(df.iloc[[-1]]['High']))
		buy_list.append(buy)
		
def test_add():
	print(buy_list)
	df2=pd.DataFrame(buy_list)
	df2.to_excel('MIDCAP150'+str(datetime.date.today())+'.xlsx')

'''
print(read_data())
