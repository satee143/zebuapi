import os

import openpyxl
from openpyxl.utils.cell import get_column_letter

global wb
os.chdir('/storage/emulated/0/Download')

wb = openpyxl.load_workbook('a.xlsx')
wb.sheetnames
sheet = wb["Sheet1"]
amountOfRows = sheet.max_row
amountOfColumns = sheet.max_column
print(amountOfRows, amountOfColumns)

for i in range(amountOfColumns):
    for k in range(amountOfRows):
        cell = sheet[get_column_letter(i + 1) + str(k + 1)]
        print(cell)

        if (str(cell.value) == "1.1.8"):
            m = str(cell)[15:18]

            sheet[m] = 'good to go'

            wb.save('a.xlsx')

        elif (str(cell.value) == "Hrs Picker"):
            m = str(cell)[15:18]
            print()
            sheet[m] = 'good to go'
            cell = 'good update'
            wb.save('a.xlsx')
        # print('matched')

'''
            newCell = "'=,"+cell[1:]
            sheet[get_column_letter(i+1)+str(k+1)]=newCell
'''
