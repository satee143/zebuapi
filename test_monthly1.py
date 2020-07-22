import plotly.graph_objects as go
import pandas as pd
import os
import nsepy
import datetime
from datetime import timedelta  
import openpyxl

global df,buy_list
global updates
buy_list=[]
updates=[]
import pytest

os.chdir('/storage/emulated/0/Download')

def read_data():

    wb = openpyxl.load_workbook('nse_quotes.xlsx')
    ws = wb['data2']
    row = ws.max_row
    l = []

    for i in range(1, row + 1):
        l1 = []
        stock_name = ws.cell(i, 1)

        l.insert(0, stock_name.value)

    print(l)
    return l

@pytest.mark.parametrize('data', read_data())
def test_create_chart(data):
	date_y=datetime.date.today() - timedelta(days=31)
	df = nsepy.get_history(symbol=data, start=date_y, end=datetime.date.today())
	#df = nsepy.get_history(symbol=data, start=datetime.date(2020,4,1), end=datetime.date.today())
	buy=[]
	df.reset_index('Date', inplace=True)
	df['Date'] = df['Date'].astype(str)
	df['Date'] = df['Date'].apply(lambda x: datetime.datetime.strptime(x, '%Y-%m-%d')) 
	df.set_index('Date', inplace=True) 
	ohlc_dict = {'Open':'first', 'High':'max', 'Low':'min', 'Close': 'last'}

	df = df.resample('1M', how=ohlc_dict).dropna(how='any')
	#df = df['Close'].resample('1M').ohlc() 
	df.reset_index('Date', inplace=True) 
	fig = go.Figure(data=[go.Ohlc(x=df['Date'], open=df['Open'],high=df['High'], low=df['Low'], close=df['Close'])])
	df['HA_Close'] = round((df['Open'] + df['High']+ df['Low'] + df['Close']) / 4, 2)
	df['HA_Open'] = (df['Open'].shift(1) +df['Open'].shift(1)) / 2
	df.iloc[0, df.columns.get_loc("HA_Open")] = (df.iloc[0]['Open'] + df.iloc[0]['Close']) / 2 
	df['HA_High'] = df[['High', 'Low', 'HA_Open', 'HA_Close']].max(axis=1) 
	df['HA_Low'] = df[['High', 'Low', 'HA_Open', 'HA_Close']].min(axis=1) 
	fig = go.Figure(data=[go.Candlestick(x=df['Date'], open=df['HA_Open'], high=df['HA_High'], low=df['HA_Low'],close=df['HA_Close'])])
		
	if (float(df.iloc[[-1]]['HA_High']) <= float(df.iloc[[-1]]['HA_Open'])): 
		if (float(df.iloc[[-1]]['HA_Low']) <= float(df.iloc[[-1]]['HA_Close'])):
			print('Sell Recommanded Price for ' + data + ' :', round(float(df.iloc[[-1]]['HA_Low']) * 98 / 100, 2))
	elif (float(df.iloc[[-1]]['HA_Low']) >= float(df.iloc[[-1]]['HA_Open'])):
		if ((float(df.iloc[[-1]]['HA_High']) >= float(df.iloc[[-1]]['HA_Close']))): 
			buy.append(data) 
			buy.append(round(float(df.iloc[[-1]]['HA_High']) * 102 / 100, 2)) 
			print('Buy Recommanded Price for ' + data + ':', round(float(df.iloc[[-1]]['HA_High']) * 102 / 100, 2))
			os.chdir('/storage/emulated/0/recom') 
			df[['Date', 'HA_Open', 'HA_High', 'HA_Low', 'HA_Close']].to_excel(data+'.xlsx') 
			fig.write_html(data+'.html') 
			buy_list.append(buy)
			#fig.show()
    
def test_add():
	os.chdir('/storage/emulated/0/Download')

	book=openpyxl.load_workbook('monthly_recom.xlsx')
	ws= book.create_sheet(str(datetime.date.today().strftime('%b'))+'_Recom')
	for row in buy_list:
		ws.append(row)
	
	book.save('monthly_recom.xlsx')
