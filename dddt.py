import os

import openpyxl

os.chdir('/storage/emulated/0/Download')
m = open('/storage/emulated/0/Download/x.txt', 'w')


def read_data():
    wb = openpyxl.load_workbook('/storage/emulated/0/Download/data.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    l = []
    for i in range(1, row + 1):
        l1 = []
        username = ws.cell(i, 1)
        l1.insert(0, username.value)
        l.insert(i - 1, l1)
    print(l, file=m)
    return l


read_data()
