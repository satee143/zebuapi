# imlint:disable=W0311
import os

import nsepy
import openpyxl
import pytest

import datetime
from datetime import timedelta

f_list = []

os.chdir('/storage/emulated/0/Download')


def read_data():
    wb = openpyxl.load_workbook('quote.xlsx')

    # ws = wb[str(datetime.date.today())]
    ws = wb['2020-05-29']
    row = ws.max_row
    l = []
    for i in range(1, row + 1):
        l1 = []
        stock_name = ws.cell(i, 1)
        stock_type = ws.cell(i, 2)
        stock_price = ws.cell(i, 3)
        l1.insert(0, stock_name.value)
        l1.insert(1, stock_type.value)
        l1.insert(2, stock_price.value)
        l.insert(i - 1, l1)

    print(l)
    return l


@pytest.mark.parametrize('data', read_data())
def test_create_chart(data):
    date_y = datetime.date.today() - timedelta(days=4)
    df = nsepy.get_history(symbol=data[0], start=date_y, end=datetime.date.today())
    t_list = []

    if data[1] == 'Buy':
        if float(df.iloc[[-1]]['High']) >= float(data[2]):
            profit = float(df.iloc[[-1]]['High']) - float(data[2])
            profit_percent = round((profit / float(data[2])) * 100, 2)
            t_list.append(str(datetime.date.today()))
            t_list.append(data[0])
            t_list.append(data[1])
            t_list.append(data[2])
            t_list.append(float(df.iloc[[-1]]['High']))
            t_list.append(profit_percent)
            f_list.append(t_list)


    elif data[1] == 'Sell':
        if float(df.iloc[[-1]]['Low']) <= float(data[2]):
            profit = float(data[2]) - float(df.iloc[[-1]]['Low'])
            profit_percent = round((profit / float(data[2])) * 100, 2)
            t_list.append(str(datetime.date.today()))
            t_list.append(data[0])
            t_list.append(data[1])
            t_list.append(data[2])
            t_list.append(float(df.iloc[[-1]]['Low']))
            t_list.append(profit_percent)
            f_list.append(t_list)


def test_to_excel():
    ''' To Copy the result set to Excel'''
    book = openpyxl.load_workbook('ledger.xlsx')

    sheet = book.active
    for row in f_list:
        sheet.append(row)

    book.save('ledger.xlsx')
    read_data()
